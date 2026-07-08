"""
===========================================================
AulaMind Enterprise 3.0
services/curriculum_loader.py
-----------------------------------------------------------

Servicio encargado de importar automáticamente el
Currículum Nacional de Chile.

Formatos soportados:

• Excel (.xlsx)
• CSV
• JSON

Importa:

• Cursos
• Asignaturas
• Unidades
• Objetivos de Aprendizaje (OA)

Autor:
Biotecno Chile
===========================================================
"""

from pathlib import Path
import json
import csv

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database.session import SessionLocal

from models.course import Course
from models.subject import Subject
from models.unit import Unit
from models.learning_objective import LearningObjective


class CurriculumLoader:
    """
    Servicio encargado de importar el Currículum
    Nacional de Chile.
    """

    # =====================================================
    # CONSTRUCTOR
    # =====================================================

    def __init__(self, db: Session = None):

        self.db = db or SessionLocal()

        self.base_path = Path("data/curriculum")

        self.supported_extensions = [

            ".xlsx",

            ".csv",

            ".json"

        ]

        self.reset_statistics()

    # =====================================================
    # CONFIGURACIÓN
    # =====================================================

    def set_base_path(self, path):

        self.base_path = Path(path)

    def get_base_path(self):

        return self.base_path

    def supported_formats(self):

        return self.supported_extensions

    # =====================================================
    # ESTADÍSTICAS
    # =====================================================

    def reset_statistics(self):

        self.statistics = {

            "files": 0,

            "courses": 0,

            "subjects": 0,

            "units": 0,

            "learning_objectives": 0,

            "inserted": 0,

            "updated": 0,

            "ignored": 0,

            "errors": 0

        }

    def get_statistics(self):

        return self.statistics

    def print_statistics(self):

        print()

        print("=" * 60)

        print("CURRICULUM IMPORT REPORT")

        print("=" * 60)

        for key, value in self.statistics.items():

            print(f"{key:25} : {value}")

        print("=" * 60)

    # =====================================================
    # VALIDACIONES
    # =====================================================

    def file_exists(self, file_path):

        return Path(file_path).exists()

    def validate_extension(self, file_path):

        ext = Path(file_path).suffix.lower()

        return ext in self.supported_extensions

    def validate_file(self, file_path):

        if not self.file_exists(file_path):

            raise FileNotFoundError(

                f"No existe el archivo:\n{file_path}"

            )

        if not self.validate_extension(file_path):

            raise Exception(

                "Formato no soportado.\n"

                "Utilice Excel (.xlsx), CSV o JSON."

            )

        return True

    def validate_record(self, record):

        required = [

            "course",

            "subject",

            "unit",

            "oa",

            "title"

        ]

        missing = []

        for field in required:

            if field not in record:

                missing.append(field)

                continue

            value = str(record[field]).strip()

            if value == "":

                missing.append(field)

        if missing:

            raise Exception(

                "Campos obligatorios faltantes: "

                + ", ".join(missing)

            )

        return True

    # =====================================================
    # NORMALIZACIÓN
    # =====================================================

    def normalize(self, value):

        if value is None:

            return ""

        return str(value).strip()

    def normalize_record(self, record):

        normalized = {}

        for key, value in record.items():

            normalized[key.strip().lower()] = self.normalize(value)

        return normalized

    # =====================================================
    # UTILIDADES
    # =====================================================

    def close(self):

        if self.db:

            self.db.close()

    def rollback(self):

        self.db.rollback()

    def commit(self):

        self.db.commit()

    def flush(self):

        self.db.flush()
            # =====================================================
    # IMPORTACIÓN DESDE EXCEL
    # =====================================================

    def load_excel(self, file_path):
        """
        Carga un archivo Excel del Currículum Nacional.

        Retorna una lista de registros normalizados.
        """

        self.validate_file(file_path)

        workbook = load_workbook(

            filename=file_path,

            data_only=True

        )

        self.statistics["files"] += 1

        records = []

        for sheet_name in workbook.sheetnames:

            worksheet = workbook[sheet_name]

            print(f"Leyendo hoja: {sheet_name}")

            sheet_records = self.read_sheet(worksheet)

            records.extend(sheet_records)

        print(f"Total registros encontrados: {len(records)}")

        return records

    # =====================================================
    # LECTURA DE UNA HOJA
    # =====================================================

    def read_sheet(self, worksheet):
        """
        Lee una hoja del Excel.

        La primera fila debe contener los encabezados.

        Retorna una lista de diccionarios.
        """

        rows = list(

            worksheet.iter_rows(values_only=True)

        )

        if len(rows) <= 1:

            return []

        headers = []

        for header in rows[0]:

            if header is None:

                headers.append("")

            else:

                headers.append(

                    str(header).strip().lower()

                )

        records = []

        for row in rows[1:]:

            if row is None:

                continue

            record = {}

            for index, value in enumerate(row):

                if index >= len(headers):

                    continue

                key = headers[index]

                if key == "":

                    continue

                record[key] = value

            if len(record) == 0:

                continue

            try:

                normalized = self.normalize_excel_record(record)

                self.validate_record(normalized)

                records.append(normalized)

            except Exception as ex:

                self.statistics["errors"] += 1

                print(f"Registro ignorado: {ex}")

        return records

    # =====================================================
    # NORMALIZACIÓN
    # =====================================================

    def normalize_excel_record(self, record):
        """
        Convierte una fila del Excel
        al formato oficial de AulaMind.
        """

        data = {}

        data["course"] = self.normalize(

            record.get("course")

            or record.get("curso")

        )

        data["subject"] = self.normalize(

            record.get("subject")

            or record.get("asignatura")

        )

        data["unit"] = self.normalize(

            record.get("unit")

            or record.get("unidad")

        )

        data["oa"] = self.normalize(

            record.get("oa")

            or record.get("objetivo")

            or record.get("codigo")

        )

        data["title"] = self.normalize(

            record.get("title")

            or record.get("titulo")

        )

        data["description"] = self.normalize(

            record.get("description")

            or record.get("descripcion")

        )

        data["indicators"] = self.normalize(

            record.get("indicators")

            or record.get("indicadores")

        )

        data["skills"] = self.normalize(

            record.get("skills")

            or record.get("habilidades")

        )

        data["attitudes"] = self.normalize(

            record.get("attitudes")

            or record.get("actitudes")

        )

        data["semester"] = self.to_integer(

            record.get("semester")

            or record.get("semestre"),

            default=1

        )

        data["weeks"] = self.to_integer(

            record.get("weeks")

            or record.get("semanas"),

            default=4

        )

        data["hours"] = self.to_integer(

            record.get("hours")

            or record.get("horas"),

            default=24

        )

        return data

    # =====================================================
    # CONVERSIONES
    # =====================================================

    def to_integer(self, value, default=0):
        """
        Convierte cualquier valor a entero.
        """

        if value is None:

            return default

        if value == "":

            return default

        try:

            return int(value)

        except Exception:

            return default

    # =====================================================
    # PREVISUALIZACIÓN
    # =====================================================

    def preview_excel(self, file_path, limit=10):
        """
        Devuelve los primeros registros
        sin insertarlos en la base de datos.
        """

        records = self.load_excel(file_path)

        return records[:limit]

    # =====================================================
    # INFORMACIÓN DEL ARCHIVO
    # =====================================================

    def excel_information(self, file_path):
        """
        Obtiene información básica
        del archivo Excel.
        """

        self.validate_file(file_path)

        workbook = load_workbook(

            filename=file_path,

            data_only=True

        )

        info = {

            "file": str(file_path),

            "worksheets": workbook.sheetnames,

            "total_sheets": len(workbook.sheetnames)

        }

        return info
            # =====================================================
    # IMPORTACIÓN DESDE CSV
    # =====================================================

    def load_csv(self, file_path):
        """
        Importa un archivo CSV del Currículum Nacional.
        """

        self.validate_file(file_path)

        self.statistics["files"] += 1

        records = []

        with open(
            file_path,
            "r",
            encoding="utf-8-sig",
            newline=""
        ) as csvfile:

            reader = csv.DictReader(csvfile)

            for row in reader:

                try:

                    record = self.normalize_csv_record(row)

                    self.validate_record(record)

                    records.append(record)

                except Exception as ex:

                    self.statistics["errors"] += 1

                    print(f"Registro CSV ignorado: {ex}")

        print(f"CSV importado: {len(records)} registros")

        return records

    # =====================================================
    # NORMALIZAR CSV
    # =====================================================

    def normalize_csv_record(self, row):

        normalized = {}

        for key, value in row.items():

            normalized[str(key).strip().lower()] = self.normalize(value)

        return self.normalize_excel_record(normalized)

    # =====================================================
    # IMPORTACIÓN DESDE JSON
    # =====================================================

    def load_json(self, file_path):
        """
        Importa un archivo JSON.
        """

        self.validate_file(file_path)

        self.statistics["files"] += 1

        with open(
            file_path,
            "r",
            encoding="utf-8"
        ) as file:

            data = json.load(file)

        records = []

        if isinstance(data, dict):

            data = data.get("learning_objectives", [])

        for item in data:

            try:

                record = self.normalize_json_record(item)

                self.validate_record(record)

                records.append(record)

            except Exception as ex:

                self.statistics["errors"] += 1

                print(f"Registro JSON ignorado: {ex}")

        print(f"JSON importado: {len(records)} registros")

        return records

    # =====================================================
    # NORMALIZAR JSON
    # =====================================================

    def normalize_json_record(self, record):

        normalized = {}

        normalized["course"] = self.normalize(

            record.get("course")

        )

        normalized["subject"] = self.normalize(

            record.get("subject")

        )

        normalized["unit"] = self.normalize(

            record.get("unit")

        )

        normalized["oa"] = self.normalize(

            record.get("oa")

        )

        normalized["title"] = self.normalize(

            record.get("title")

        )

        normalized["description"] = self.normalize(

            record.get("description")

        )

        normalized["indicators"] = self.normalize(

            record.get("indicators")

        )

        normalized["skills"] = self.normalize(

            record.get("skills")

        )

        normalized["attitudes"] = self.normalize(

            record.get("attitudes")

        )

        normalized["semester"] = self.to_integer(

            record.get("semester"),

            default=1

        )

        normalized["weeks"] = self.to_integer(

            record.get("weeks"),

            default=4

        )

        normalized["hours"] = self.to_integer(

            record.get("hours"),

            default=24

        )

        return normalized

    # =====================================================
    # DETECTAR FORMATO
    # =====================================================

    def load(self, file_path):
        """
        Detecta automáticamente el formato
        del archivo e importa los datos.
        """

        extension = Path(file_path).suffix.lower()

        if extension == ".xlsx":

            return self.load_excel(file_path)

        if extension == ".csv":

            return self.load_csv(file_path)

        if extension == ".json":

            return self.load_json(file_path)

        raise Exception(

            f"Formato no soportado: {extension}"

        )

    # =====================================================
    # PREVISUALIZAR
    # =====================================================

    def preview(self, file_path, limit=20):
        """
        Devuelve una vista previa del archivo
        sin insertar datos.
        """

        records = self.load(file_path)

        return records[:limit]

    # =====================================================
    # VALIDACIÓN GENERAL
    # =====================================================

    def validate_dataset(self, records):
        """
        Valida un conjunto completo
        de registros.
        """

        valid = 0

        invalid = 0

        for record in records:

            try:

                self.validate_record(record)

                valid += 1

            except Exception:

                invalid += 1

        return {

            "total": len(records),

            "valid": valid,

            "invalid": invalid

        }
            # =====================================================
    # IMPORTAR DATASET
    # =====================================================

    def import_dataset(self, records):
        """
        Importa un conjunto de registros
        a la base de datos.
        """

        print()

        print("=" * 60)
        print("IMPORTANDO CURRÍCULUM")
        print("=" * 60)

        try:

            for record in records:

                self.import_record(record)

            self.commit()

            print()

            print("Importación finalizada correctamente.")

            return True

        except Exception as ex:

            self.rollback()

            print()

            print("ERROR:", ex)

            raise ex

    # =====================================================
    # IMPORTAR UN REGISTRO
    # =====================================================

    def import_record(self, record):

        course = self.get_or_create_course(record)

        subject = self.get_or_create_subject(
            course,
            record
        )

        unit = self.get_or_create_unit(
            course,
            subject,
            record
        )

        self.get_or_create_learning_objective(
            course,
            subject,
            unit,
            record
        )

    # =====================================================
    # COURSE
    # =====================================================

    def get_or_create_course(self, record):

        name = record["course"]

        course = self.db.query(Course).filter(

            Course.name == name

        ).first()

        if course:

            return course

        course = Course(

            name=name,

            short_name=name,

            level="Básica",

            order=1

        )

        self.db.add(course)

        self.flush()

        self.statistics["courses"] += 1
        self.statistics["inserted"] += 1

        return course

    # =====================================================
    # SUBJECT
    # =====================================================

    def get_or_create_subject(

        self,

        course,

        record

    ):

        name = record["subject"]

        subject = self.db.query(Subject).filter(

            Subject.course_id == course.id,

            Subject.name == name

        ).first()

        if subject:

            return subject

        subject = Subject(

            course_id=course.id,

            name=name,

            short_name=name,

            code=name[:10].upper(),

            description="",

            order=1

        )

        self.db.add(subject)

        self.flush()

        self.statistics["subjects"] += 1
        self.statistics["inserted"] += 1

        return subject

    # =====================================================
    # UNIT
    # =====================================================

    def get_or_create_unit(

        self,

        course,

        subject,

        record

    ):

        title = record["unit"]

        unit = self.db.query(Unit).filter(

            Unit.subject_id == subject.id,

            Unit.title == title

        ).first()

        if unit:

            return unit

        number = 1

        try:

            number = int(

                ''.join(

                    filter(str.isdigit, title)

                )

            )

        except Exception:

            pass

        unit = Unit(

            course_id=course.id,

            subject_id=subject.id,

            number=number,

            title=title,

            description="",

            semester=record.get("semester", 1),

            estimated_weeks=record.get("weeks", 4),

            estimated_hours=record.get("hours", 24),

            order=number

        )

        self.db.add(unit)

        self.flush()

        self.statistics["units"] += 1
        self.statistics["inserted"] += 1

        return unit

    # =====================================================
    # LEARNING OBJECTIVE
    # =====================================================

    def get_or_create_learning_objective(

        self,

        course,

        subject,

        unit,

        record

    ):

        code = record["oa"]

        oa = self.db.query(

            LearningObjective

        ).filter(

            LearningObjective.code == code,

            LearningObjective.subject_id == subject.id

        ).first()

        if oa:

            oa.title = record["title"]

            oa.description = record["description"]

            oa.indicators = record["indicators"]

            oa.skills = record["skills"]

            oa.attitudes = record["attitudes"]

            oa.estimated_hours = record.get(

                "hours",

                6

            )

            self.statistics["updated"] += 1

            return oa

        oa = LearningObjective(

            course_id=course.id,

            subject_id=subject.id,

            unit_id=unit.id,

            code=code,

            title=record["title"],

            description=record["description"],

            indicators=record["indicators"],

            skills=record["skills"],

            attitudes=record["attitudes"],

            estimated_hours=record.get(

                "hours",

                6

            ),

            minimum_classes=1,

            bloom_level="Comprender",

            priority=1,

            is_transversal=False,

            is_active=True

        )

        self.db.add(oa)

        self.flush()

        self.statistics["learning_objectives"] += 1
        self.statistics["inserted"] += 1

        return oa
            # =====================================================
    # IMPORTAR ARCHIVO COMPLETO
    # =====================================================

    def import_file(self, file_path):
        """
        Método principal.

        Detecta automáticamente el formato,
        importa el archivo y lo almacena
        en la base de datos.
        """

        self.reset_statistics()

        try:

            records = self.load(file_path)

            validation = self.validate_dataset(records)

            print()
            print("=" * 60)
            print("VALIDACIÓN DEL DATASET")
            print("=" * 60)
            print(f"Total registros : {validation['total']}")
            print(f"Válidos         : {validation['valid']}")
            print(f"Inválidos       : {validation['invalid']}")
            print("=" * 60)
            print()

            if validation["valid"] == 0:

                raise Exception(
                    "No existen registros válidos."
                )

            self.import_dataset(records)

            self.commit()

            self.print_statistics()

            return True

        except Exception as ex:

            self.rollback()

            print()

            print("ERROR DURANTE LA IMPORTACIÓN")

            print(str(ex))

            raise ex

        finally:

            self.close()

    # =====================================================
    # LOG
    # =====================================================

    def log(self, message):

        print(

            f"[CurriculumLoader] {message}"

        )

    def warning(self, message):

        print(

            f"[WARNING] {message}"

        )

    def error(self, message):

        print(

            f"[ERROR] {message}"

        )

    # =====================================================
    # RESUMEN
    # =====================================================

    def summary(self):
        """
        Devuelve un resumen
        de la importación.
        """

        return {

            "files": self.statistics["files"],

            "courses": self.statistics["courses"],

            "subjects": self.statistics["subjects"],

            "units": self.statistics["units"],

            "learning_objectives":
                self.statistics["learning_objectives"],

            "inserted":
                self.statistics["inserted"],

            "updated":
                self.statistics["updated"],

            "ignored":
                self.statistics["ignored"],

            "errors":
                self.statistics["errors"]

        }

    # =====================================================
    # EXPORTAR REPORTE
    # =====================================================

    def export_report(self):

        report = []

        report.append("=" * 60)

        report.append("AULAMIND CURRICULUM REPORT")

        report.append("=" * 60)

        for key, value in self.statistics.items():

            report.append(

                f"{key:<30} {value}"

            )

        report.append("=" * 60)

        return "\n".join(report)

    # =====================================================
    # LIMPIAR BASE CURRICULAR
    # =====================================================

    def clear_database(self):
        """
        Elimina toda la información
        curricular.

        Útil durante desarrollo.
        """

        print()

        print("Eliminando Base Curricular...")

        self.db.query(

            LearningObjective

        ).delete()

        self.db.query(

            Unit

        ).delete()

        self.db.query(

            Subject

        ).delete()

        self.db.query(

            Course

        ).delete()

        self.commit()

        print("Base Curricular eliminada.")

    # =====================================================
    # REIMPORTAR
    # =====================================================

    def rebuild_database(self, file_path):
        """
        Borra completamente
        la base curricular
        y vuelve a importarla.
        """

        self.clear_database()

        self.import_file(file_path)

    # =====================================================
    # INFORMACIÓN
    # =====================================================

    def version(self):

        return "1.0.0"

    def author(self):

        return "Biotecno Chile"

    def supported_files(self):

        return [

            "*.xlsx",

            "*.csv",

            "*.json"

        ]

    # =====================================================
    # FIN
    # =====================================================